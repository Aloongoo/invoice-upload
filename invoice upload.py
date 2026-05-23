import os
import base64
from io import BytesIO
import streamlit as st
from PIL import Image, ImageEnhance, ImageFilter, ImageOps
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from google import genai

# 🌟 1. 锁死你的尊贵付费版 Gemini Key（对前端用户隐形，安全无漏）
os.environ["GEMINI_API_KEY"] = "AIzaSyDFfPJBuWgBD0Yr8k7bHU8dKGdz2b50SzU"

# 🌟 2. 网页基础配置 (高级商用 SaaS 外观)
st.set_page_config(
    page_title="Papa Sushi 财务智能发票系统", 
    page_icon="🧾", 
    layout="wide"
)

# ─── 侧边栏：商户权限与状态 ───
st.sidebar.title("🍱 控制台")
st.sidebar.write("欢迎使用 Papa Sushi 财务自动化托管方案。")

license_type = st.sidebar.selectbox("🔑 您的商户版本:", ["标准企业版 (Premium)", "体验试用版"])
st.sidebar.success(f"🟢 系统状态: 付费通道已连接")

st.sidebar.markdown("---")
# ♻️ 内存自净按钮：随手一解，1GB服务器绝不卡死
if st.sidebar.button("♻️ 强制销毁数据并清空内存", type="secondary"):
    st.session_state.clear()
    st.cache_data.clear()
    st.sidebar.success("💥 内存已彻底清空！历史单据已被物理销毁。")
    st.rerun()

st.sidebar.markdown("---")
st.sidebar.info("📢 **商户机密安全**\n\n本系统采用纯内存流技术，AI 审计完毕并交付下载后，图片不留存服务器，保障您的商业机密安全。")

# ─── 主界面 ───
st.title("🧾 智账宝 · 全自动发票核销与 A4 排版系统")
st.write("一键批量提取消费数据、生成 Excel 报表，并全自动将小票排版为直立 A4 田字格 PDF。")

# --- 核心图像处理（抹掉阴影、自动纠正旋转） ---
def load_image_safely(uploaded_file):
    try:
        file_bytes = uploaded_file.read()
        if uploaded_file.name.lower().endswith('.heic'):
            from pillow_heif import register_heif_opener
            register_heif_opener()
        return Image.open(BytesIO(file_bytes))
    except Exception as e:
        st.error(f"⚠️ 无法读取图片 [{uploaded_file.name}]，可能文件损坏。原因: {e}")
        return None

def clean_image(img):
    try:
        img = ImageOps.exif_transpose(img)
        if img.width > img.height:
            img = img.rotate(-90, expand=True)
        img_gray = img.convert('L')
        bg = img_gray.filter(ImageFilter.GaussianBlur(radius=20))
        res = np.array(img_gray, dtype=float) / np.array(bg, dtype=float) * 255
        res = np.clip(res, 0, 255).astype(np.uint8)
        cleaned_img = Image.fromarray(res)
        return ImageEnhance.Contrast(cleaned_img).enhance(2.0).convert('1')
    except Exception:
        return img.convert('RGB')

# --- 核心 AI 审计大脑 ---
def extract_invoice_data_via_gemini(img, file_name):
    try:
        client = genai.Client()
        img = ImageOps.exif_transpose(img)
        
        prompt = """
        你是一个精明的财务审计专家。请仔细阅读这张发票/收据图片，提取以下信息。
        请严格按照下面的格式返回数据，不要有任何多余的解释或废话：
        商家: [这里写商店名字]
        日期: [这里写日期，统一格式为 YYYY-MM-DD，如果找不到写 9999-12-31]
        单号: [这里写发票号码或收据号码。如果实在找不到，写 找不到单号]
        总金额: [这里只写最终付的数字，不要货币符号]
        明细: [简要列出购买的核心物品或服务，用逗号隔开]
        """
        
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=[img, prompt]
        )
        
        lines = response.text.strip().split('\n')
        data = {"商家": "未知商家", "日期": "9999-12-31", "单号": "找不到单号", "总金额": "0.00", "明细": "无", "原始照片名": file_name}
        
        for line in lines:
            if "商家:" in line or "商家：" in line:
                data["商家"] = line.split(":", 1)[-1].split("：", 1)[-1].strip()
            elif "日期:" in line or "日期：" in line:
                data["日期"] = line.split(":", 1)[-1].split("：", 1)[-1].strip()
            elif "单号:" in line or "单号：" in line:
                data["单号"] = line.split(":", 1)[-1].split("：", 1)[-1].strip()
            elif "总金额:" in line or "总金额：" in line:
                data["总金额"] = line.split(":", 1)[-1].split("：", 1)[-1].strip()
            elif "明细:" in line or "明细：" in line:
                data["明细"] = line.split(":", 1)[-1].split("：", 1)[-1].strip()
        
        for char in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
            data["商家"] = data["商家"].replace(char, '')
            
        return data
    except Exception as e:
        return {"商家": "解析异常", "日期": "9999-12-31", "单号": "失败", "总金额": "0.00", "明细": f"原因: {str(e)}", "原始照片名": file_name}

# --- 网页核心文件上传区 ---
uploaded_files = st.file_uploader("📂 请批量拖入发票照片 (支持多选，支持手机拍的 JPG, PNG, HEIC 格式)", type=["jpg", "jpeg", "png", "heic"], accept_multiple_files=True)

if uploaded_files:
    st.info(f"📁 成功捕获到该商户的 {len(uploaded_files)} 张待核销单据。")
    
    if st.button("🚀 启动全自动智能核销排版流水线", type="primary"):
        all_records = []
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        for idx, uploaded_file in enumerate(uploaded_files):
            status_text.text(f"⏳ 正在流水线作业 ({idx+1}/{len(uploaded_files)}): {uploaded_file.name}...")
            
            img = load_image_safely(uploaded_file)
            if img:
                record = extract_invoice_data_via_gemini(img, uploaded_file.name)
                record["_img_obj"] = img 
                all_records.append(record)
            
            progress_bar.progress((idx + 1) / len(uploaded_files))
        
        status_text.text("✨ AI 商业审计完成！正在为您重新按日期排序并进行 A4 拼图排版...")
        all_records.sort(key=lambda x: x["日期"])
        
        st.subheader("📊 财务登账预览总表 (已按日期自动排序)")
        display_data = [{k: v for k, v in r.items() if not k.startswith('_')} for r in all_records]
        st.dataframe(display_data, use_container_width=True)
        
        # --- 后台构建商用级 Excel ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "发票核销总表"
        headers = ["消费日期", "商家名称", "发票/收据单号", "最终总金额", "对应排版文件", "A4纸位置", "购买物品明细", "原始照片名"]
        ws.append(headers)
        
        for col_idx in range(1, 9):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        merchant_buckets = {}
        for r in all_records:
            m = r["商家"]
            if m not in merchant_buckets: merchant_buckets[m] = []
            merchant_buckets[m].append(r)
            
        a4_w, a4_h = 2480, 3508
        target_w, target_h = int(a4_w / 2) - 40, int(a4_h / 2) - 60
        positions = [(20, 40), (int(a4_w / 2) + 20, 40), (20, int(a4_h / 2) + 20), (int(a4_w / 2) + 20, int(a4_h / 2) + 20)]
        pos_names = ["左上角", "右上角", "左下角", "右下角"]
        
        pdf_bytes_dict = {}
        
        for merchant, records in merchant_buckets.items():
            page_number = 1
            while len(records) > 0:
                current_batch = records[:4]
                records = records[4:]
                
                a4_canvas = Image.new('RGB', (a4_w, a4_h), 'white')
                output_pdf_name = f"A4版_{merchant}_第{page_number}页.pdf"
                
                for index, info in enumerate(current_batch):
                    cleaned_img = clean_image(info["_img_obj"])
                    cleaned_img.thumbnail((target_w, target_h))
                    a4_canvas.paste(cleaned_img, positions[index])
                    
                    display_date = "未知日期" if info["日期"] == "9999-12-31" else info["日期"]
                    ws.append([
                        display_date, info["商家"], info["单号"], 
                        float(info["总金额"]) if info["总金额"].replace('.', '', 1).isdigit() else info["总金额"], 
                        output_pdf_name, pos_names[index], info["明细"], info["原始照片名"]
                    ])
                
                pdf_buffer = BytesIO()
                a4_canvas.save(pdf_buffer, "PDF", resolution=300.0, optimize=True)
                pdf_bytes_dict[output_pdf_name] = pdf_buffer.getvalue()
                page_number += 1
                
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 13)
            
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        
        # --- 成果交付中心 ---
        st.subheader("📥 成果交付中心")
        c1, c2 = st.columns(2)
        with c1:
            st.download_button(
                label="🟢 一键下载：商户财务记账总表 (Excel)",
                data=excel_buffer.getvalue(),
                file_name="发票财务记账总表.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        st.write("📄 **排版好的 A4 报税打印件 (点击即可无损下载)：**")
        for pdf_name, bytes_data in pdf_bytes_dict.items():
            st.download_button(label=f"⬇️ 导出 {pdf_name}", data=bytes_data, file_name=pdf_name, mime="application/pdf")
            
        st.success("🎉 全自动商户报账流水线作业圆满成功！")